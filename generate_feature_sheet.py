#!/usr/bin/env python3
"""
LightNVR Feature Analysis Excel Sheet Generator
Generates a comprehensive feature breakdown with status tracking.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

# ── Colour palette ──────────────────────────────────────────────────
FILL_HEADER  = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
FILL_SECTION = PatternFill(start_color="374151", end_color="374151", fill_type="solid")
FILL_GREEN   = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")  # Working / Done
FILL_RED     = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")  # Not Working / Issue
FILL_YELLOW  = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")  # Testing / Partial
FILL_BLUE    = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")   # Info / Needs Review
FILL_GRAY    = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")   # Alt row
FILL_WHITE   = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

FONT_HEADER  = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
FONT_SECTION = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
FONT_NORMAL  = Font(name="Calibri", size=10)
FONT_BOLD    = Font(name="Calibri", bold=True, size=10)

THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)

ALIGNMENT_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGNMENT_LEFT   = Alignment(horizontal="left", vertical="center", wrap_text=True)

# ── Status helpers ──────────────────────────────────────────────────
STATUS_MAP = {
    "Working":      FILL_GREEN,
    "Not Working":  FILL_RED,
    "Issue":        FILL_RED,
    "Testing":      FILL_YELLOW,
    "Partial":      FILL_YELLOW,
    "Done":         FILL_GREEN,
    "Needs Review": FILL_BLUE,
}

FONT_STATUS = {
    "Working":      Font(name="Calibri", bold=True, size=10, color="065F46"),
    "Not Working":  Font(name="Calibri", bold=True, size=10, color="991B1B"),
    "Issue":        Font(name="Calibri", bold=True, size=10, color="991B1B"),
    "Testing":      Font(name="Calibri", bold=True, size=10, color="92400E"),
    "Partial":      Font(name="Calibri", bold=True, size=10, color="92400E"),
    "Done":         Font(name="Calibri", bold=True, size=10, color="065F46"),
    "Needs Review": Font(name="Calibri", bold=True, size=10, color="1E40AF"),
}

# ── Feature data ────────────────────────────────────────────────────
# (Category, Feature, Sub-Feature/Detail, Status, Test Status, Notes)
FEATURES = [
    # ─── Authentication & Authorization ──────────────────────────────
    ("Authentication & Authorization", "Login System", "Username/password login form with session management", "Working", "Done", "HTTP Basic Auth via Mongoose server"),
    ("Authentication & Authorization", "Session Validation", "JWT/session-based auth with role checking", "Working", "Done", "validateSession() utility used across all pages"),
    ("Authentication & Authorization", "Role-Based Access Control", "Admin/Viewer roles with permission enforcement", "Working", "Done", "Admin: full access; Viewer: appearance settings only"),
    ("Authentication & Authorization", "User Management (CRUD)", "Add, edit, delete users via admin panel", "Working", "Done", "UsersView with AddUserModal, EditUserModal, DeleteUserModal"),
    ("Authentication & Authorization", "API Key Management", "Generate/revoke API keys per user", "Working", "Done", "ApiKeyModal component"),
    ("Authentication & Authorization", "Auth Toggle", "Enable/disable authentication from settings", "Working", "Done", "Setting: web_auth_enabled"),

    # ─── Live Streaming ──────────────────────────────────────────────
    ("Live Streaming", "WebRTC Live View", "Ultra-low latency live streaming via go2rtc", "Working", "Done", "WebRTCView + WebRTCVideoCell components"),
    ("Live Streaming", "HLS Live View", "Adaptive bitrate HLS streaming fallback", "Working", "Done", "LiveView + HLSVideoCell components"),
    ("Live Streaming", "WebRTC/HLS Toggle", "Switch between WebRTC and HLS modes", "Working", "Done", "Setting: webrtc_disabled; toggle in UI"),
    ("Live Streaming", "STUN/ICE NAT Traversal", "Automatic firewall traversal for WebRTC", "Working", "Done", "Configured in go2rtc.yaml"),
    ("Live Streaming", "Multi-Stream Grid Layout", "View multiple camera streams simultaneously", "Working", "Done", "Grid layout with responsive columns"),
    ("Live Streaming", "Fullscreen Mode", "Per-stream fullscreen toggle", "Working", "Done", "FullscreenManager component"),
    ("Live Streaming", "Snapshot Capture", "Capture still image from live stream", "Working", "Done", "SnapshotManager + canvas capture"),
    ("Live Streaming", "Snapshot Download", "Download captured snapshot as JPEG", "Working", "Done", "Blob-based download with auto-naming"),
    ("Live Streaming", "Stream Status Indicators", "Show connected/disconnected/error per stream", "Working", "Done", "Status badges in stream cards"),

    # ─── Stream Management ───────────────────────────────────────────
    ("Stream Management", "Add Stream", "Add new RTSP camera stream via modal form", "Working", "Done", "StreamConfigModal with full field support"),
    ("Stream Management", "Edit Stream", "Modify existing stream configuration", "Working", "Done", "StreamConfigModal in edit mode"),
    ("Stream Management", "Delete Stream", "Permanently remove stream and related data", "Working", "Done", "StreamDeleteModal with confirmation"),
    ("Stream Management", "Enable/Disable Stream", "Soft toggle stream on/off", "Working", "Done", "Toggle in stream card"),
    ("Stream Management", "Test Connection", "Verify RTSP URL connectivity", "Working", "Done", "api_handlers_streams_test.c backend"),
    ("Stream Management", "Stream Config Accordion", "Collapsible sections in config modal", "Working", "Done", "AccordionSection component in StreamConfigModal"),
    ("Stream Management", "Resolution Config", "Set width, height per stream", "Working", "Done", "1080p max, configurable per stream"),
    ("Stream Management", "FPS Control", "Configurable 1–15 FPS per stream", "Working", "Done", "Frame rate per stream"),
    ("Stream Management", "Codec Selection", "H.264 (primary), H.265 support", "Working", "Done", "Codec field in stream config"),
    ("Stream Management", "Priority Setting", "Stream priority for resource allocation", "Working", "Done", "Priority field on stream config"),
    ("Stream Management", "Recording Toggle", "Enable/disable recording per stream", "Working", "Done", "record flag per stream"),
    ("Stream Management", "Segment Duration Config", "Configure recording segment length", "Working", "Done", "Segment duration in stream config"),
    ("Stream Management", "Output Format Selection", "Choose MP4 or MKV container", "Working", "Done", "Format selection in stream config"),

    # ─── ONVIF Integration ───────────────────────────────────────────
    ("ONVIF Integration", "ONVIF Device Discovery", "Auto-discover cameras on network", "Working", "Done", "onvif_discovery.c with probe/response handling"),
    ("ONVIF Integration", "ONVIF Credentials", "Username/password for ONVIF devices", "Working", "Done", "Credential modal in StreamsView"),
    ("ONVIF Integration", "ONVIF Profile Selection", "Select stream profile from discovered device", "Working", "Done", "Profile info from device management"),
    ("ONVIF Integration", "ONVIF Motion Detection", "Receive motion events from ONVIF cameras", "Working", "Done", "onvif_detection.c + onvif_motion_recording.c"),
    ("ONVIF Integration", "ONVIF Network Override", "Override IP/port for ONVIF responses", "Working", "Done", "Network address override implementation"),
    ("ONVIF Integration", "Test Motion Event", "Trigger simulated ONVIF motion event", "Working", "Done", "triggerTestMotionEvent in StreamsView"),

    # ─── PTZ Controls ────────────────────────────────────────────────
    ("PTZ (Pan-Tilt-Zoom)", "Pan/Tilt Control", "Directional movement buttons", "Working", "Done", "PTZControls with 8-direction pad"),
    ("PTZ (Pan-Tilt-Zoom)", "Zoom Control", "Zoom in/out buttons", "Working", "Done", "Zoom buttons in PTZ panel"),
    ("PTZ (Pan-Tilt-Zoom)", "PTZ Presets", "Go to preset positions", "Working", "Done", "gotoPreset API via ONVIF PTZ"),
    ("PTZ (Pan-Tilt-Zoom)", "PTZ Home", "Return to home position", "Working", "Done", "home() API function"),
    ("PTZ (Pan-Tilt-Zoom)", "PTZ Capabilities Check", "Auto-detect PTZ support per camera", "Working", "Done", "getCapabilities API"),
    ("PTZ (Pan-Tilt-Zoom)", "Continuous Movement", "Hold button for continuous move", "Working", "Done", "mouseDown/mouseUp/mouseLeave handlers"),

    # ─── Recording Management ────────────────────────────────────────
    ("Recording Management", "List Recordings", "Paginated recording list with filters", "Working", "Done", "RecordingsView with PaginationControls"),
    ("Recording Management", "Filter Recordings", "Filter by stream, date range, format", "Working", "Done", "FiltersSidebar + ActiveFilters components"),
    ("Recording Management", "Sort Recordings", "Sort by date, size, stream name", "Working", "Done", "sortBy function in RecordingsView"),
    ("Recording Management", "Play Recording", "In-browser video playback modal", "Working", "Done", "Video player modal with playback state"),
    ("Recording Management", "Download Recording", "Download MP4/MKV file", "Working", "Done", "Direct download link generation"),
    ("Recording Management", "Delete Single Recording", "Delete individual recording", "Working", "Done", "Delete button with confirmation"),
    ("Recording Management", "Batch Delete Recordings", "Delete multiple recordings at once", "Working", "Done", "BatchDeleteModal with progress tracking"),
    ("Recording Management", "Batch Delete by Filter", "Delete all recordings matching filter", "Working", "Done", "batchDeleteRecordingsByFilter function"),
    ("Recording Management", "Recording Protection", "Toggle protection to prevent auto-deletion", "Working", "Done", "toggleProtection in RecordingsView"),
    ("Recording Management", "Select All/None", "Checkbox to select/deselect all recordings", "Working", "Done", "toggleSelectAll function"),
    ("Recording Management", "URL-Based Filtering", "Persist filters in URL query params", "Working", "Done", "urlUtils.js for URL state management"),
    ("Recording Management", "Recording Sync", "Sync DB records with filesystem", "Working", "Done", "db_recordings_sync.c"),

    # ─── Timeline View ───────────────────────────────────────────────
    ("Timeline View", "Timeline Page", "Visual timeline of recordings per stream/date", "Working", "Done", "TimelinePage with ruler, segments, cursor"),
    ("Timeline View", "Timeline Ruler", "Hour-based ruler with markers", "Working", "Done", "TimelineRuler component"),
    ("Timeline View", "Timeline Segments", "Visual blocks showing recording spans", "Working", "Done", "TimelineSegments component with color coding"),
    ("Timeline View", "Timeline Cursor", "Draggable playhead for seeking", "Working", "Done", "TimelineCursor component"),
    ("Timeline View", "Timeline Player", "Integrated video player for timeline", "Working", "Done", "TimelinePlayer with stream-aware playback"),
    ("Timeline View", "Stream Selector", "Dropdown to select stream for timeline", "Working", "Done", "Stream selection in TimelineControls"),
    ("Timeline View", "Date Selector", "Date picker for timeline navigation", "Working", "Done", "Date input in TimelineControls"),
    ("Timeline View", "Speed Controls", "Playback speed adjustment (0.5x–4x)", "Working", "Done", "SpeedControls component"),
    ("Timeline View", "URL Param Sync", "Persist stream/date in URL", "Working", "Done", "parseUrlParams / updateUrlParams"),

    # ─── Detection & AI ──────────────────────────────────────────────
    ("Detection & AI", "Detection-Based Recording", "Record only when objects detected", "Working", "Done", "detection_integration.c"),
    ("Detection & AI", "light-object-detect API", "External ONNX/TFLite detection API integration", "Working", "Done", "api_detection.c + per-stream API URL config"),
    ("Detection & AI", "Detection Overlay", "Real-time bounding boxes on live streams", "Working", "Done", "DetectionOverlay component with canvas drawing"),
    ("Detection & AI", "Snapshot with Detections", "Capture frame including detection boxes", "Working", "Done", "takeSnapshotWithDetections function"),
    ("Detection & AI", "Detection Model Selection", "Choose between ONNX, TFLite, OpenCV, SOD, motion", "Working", "Done", "Model selection in stream config"),
    ("Detection & AI", "Detection Model Listing", "API to list available models", "Working", "Done", "api_handlers_detection_models.c"),
    ("Detection & AI", "Detection Results API", "Fetch detection results per stream", "Working", "Done", "api_handlers_detection_results.c"),
    ("Detection & AI", "Confidence Threshold", "Adjustable per-stream detection threshold", "Working", "Done", "Slider in stream config + global default in settings"),
    ("Detection & AI", "Pre-Detection Buffer", "Keep N seconds before detection event", "Working", "Done", "pre_detection_buffer.c"),
    ("Detection & AI", "Post-Detection Buffer", "Continue recording N seconds after last detection", "Working", "Done", "Configurable post-buffer in settings"),
    ("Detection & AI", "SOD Integration", "Embedded SOD library for on-device detection", "Working", "Done", "sod_detection.c + sod_integration.c"),
    ("Detection & AI", "SOD RealNet Models", "Lightweight real-time neural network models", "Working", "Done", "sod_realnet.c"),
    ("Detection & AI", "Built-in Motion Detection", "Frame-diff based motion detection (no model needed)", "Working", "Done", "motion_detection.c"),
    ("Detection & AI", "Unified Detection Thread", "Single thread managing all detection streams", "Working", "Done", "unified_detection_thread.c"),
    ("Detection & AI", "Detection Config Persistence", "Save/load detection settings per stream", "Working", "Done", "detection_config.c + db_motion_config.c"),

    # ─── Detection Zones ─────────────────────────────────────────────
    ("Detection Zones", "Visual Zone Editor", "Interactive canvas polygon drawing", "Working", "Done", "ZoneEditor component with canvas-based drawing"),
    ("Detection Zones", "Multiple Zones Per Stream", "Define unlimited detection zones", "Working", "Done", "Array of zones per stream"),
    ("Detection Zones", "Zone Class Filtering", "Filter detections by object class per zone", "Working", "Done", "Per-zone class filter configuration"),
    ("Detection Zones", "Zone Confidence Threshold", "Per-zone adjustable confidence", "Working", "Done", "Threshold per zone"),
    ("Detection Zones", "Zone Color Coding", "Color-coded zones for visual identification", "Working", "Done", "Color per zone in editor"),
    ("Detection Zones", "Zone Enable/Disable", "Toggle zones without deleting", "Working", "Done", "Enable flag per zone"),
    ("Detection Zones", "Zone Save/Load", "Persist zones via API", "Working", "Done", "api_handlers_zones.c + db_zones.c"),
    ("Detection Zones", "Zone Point Editing", "Drag vertices to reshape zones", "Working", "Done", "findPointNearCursor + drag logic in ZoneEditor"),
    ("Detection Zones", "Zone Delete", "Remove individual zones", "Working", "Done", "deleteSelectedZone function"),
    ("Detection Zones", "Camera Snapshot Background", "Load go2rtc snapshot as zone editor background", "Working", "Done", "loadSnapshot using go2rtc API"),
    ("Detection Zones", "Zone Filtering in Detection", "Filter detection results to zone boundaries", "Working", "Done", "zone_filter.c"),

    # ─── Settings Management ─────────────────────────────────────────
    ("Settings Management", "General Settings", "Log level configuration", "Working", "Done", "Error/Warning/Info/Debug levels"),
    ("Settings Management", "Storage Settings", "Storage path, max size, retention days", "Working", "Done", "SettingsView storage section"),
    ("Settings Management", "HLS Storage Path", "Separate path for HLS segments", "Working", "Done", "Optional secondary storage path"),
    ("Settings Management", "Auto-Delete Oldest", "Automatic cleanup of old recordings", "Working", "Done", "auto_delete_oldest flag"),
    ("Settings Management", "Database Path Config", "Configurable SQLite DB location", "Working", "Done", "db_path setting"),
    ("Settings Management", "Web Port Config", "Configurable HTTP port", "Working", "Done", "web_port setting (1–65535)"),
    ("Settings Management", "WebRTC Disable Toggle", "Force HLS-only mode", "Working", "Done", "webrtc_disabled setting"),
    ("Settings Management", "Memory Optimization", "Buffer size, swap file config", "Working", "Done", "Buffer size KB, swap toggle, swap size MB"),
    ("Settings Management", "Buffer Strategy", "Auto/go2rtc/HLS segment/memory/mmap strategies", "Working", "Done", "5 buffer strategy options"),
    ("Settings Management", "Detection Settings", "Models path, threshold, pre/post buffer", "Working", "Done", "Global detection defaults"),
    ("Settings Management", "Save/Load Settings", "Persist settings via API", "Working", "Done", "POST /api/settings"),
    ("Settings Management", "Role-Based Settings", "Viewers see only appearance; admins see all", "Working", "Done", "Conditional rendering based on user role"),
    ("Settings Management", "Config File (INI)", "INI-based configuration file support", "Working", "Done", "config.c with inih parser"),

    # ─── MQTT Integration ────────────────────────────────────────────
    ("MQTT Integration", "MQTT Enable/Disable", "Toggle MQTT event publishing", "Working", "Done", "mqtt_enabled setting"),
    ("MQTT Integration", "MQTT Broker Config", "Host, port, credentials configuration", "Working", "Done", "Full broker connection settings"),
    ("MQTT Integration", "MQTT TLS Support", "Encrypted MQTT connections", "Working", "Done", "mqtt_tls_enabled flag"),
    ("MQTT Integration", "MQTT Topic Prefix", "Customizable topic namespace", "Working", "Done", "Default: lightnvr/detections/<stream>"),
    ("MQTT Integration", "MQTT QoS Levels", "QoS 0/1/2 support", "Working", "Done", "Dropdown with 3 QoS levels"),
    ("MQTT Integration", "MQTT Retain Flag", "Retain last message for new subscribers", "Working", "Done", "Retain checkbox"),
    ("MQTT Integration", "MQTT Keepalive", "Connection health check interval", "Working", "Done", "Configurable 10–3600 seconds"),
    ("MQTT Integration", "MQTT Client Implementation", "Full MQTT client in C", "Working", "Done", "mqtt_client.c"),

    # ─── Theme & Appearance ──────────────────────────────────────────
    ("Theme & Appearance", "7 Color Themes", "Ocean Blue, Forest Green, Royal Purple, Sunset Rose, Golden Amber, Cool Slate, Default", "Working", "Done", "ThemeCustomizer component"),
    ("Theme & Appearance", "Dark/Light Mode Toggle", "Manual mode toggle with system preference detection", "Working", "Done", "toggleDarkMode in ThemeCustomizer"),
    ("Theme & Appearance", "Color Intensity Control", "Slider for theme brightness/contrast", "Working", "Done", "0–100% intensity slider"),
    ("Theme & Appearance", "Preset Intensity Buttons", "Quick intensity presets", "Working", "Done", "handlePresetIntensity function"),
    ("Theme & Appearance", "Theme Persistence", "Save theme preference locally", "Working", "Done", "LocalStorage-based persistence"),

    # ─── System Information ──────────────────────────────────────────
    ("System Information", "System Info Dashboard", "CPU, memory, storage, uptime display", "Working", "Done", "SystemView + SystemInfo component"),
    ("System Information", "Memory & Storage Details", "Detailed memory/storage breakdown", "Working", "Done", "MemoryStorage component"),
    ("System Information", "Stream Storage Breakdown", "Per-stream storage usage", "Working", "Done", "StreamStorage component"),
    ("System Information", "Network Info", "Network interface information", "Working", "Done", "NetworkInfo component"),
    ("System Information", "Active Streams Info", "Count and status of active streams", "Working", "Done", "StreamsInfo component"),
    ("System Information", "System Logs Viewer", "Real-time log viewing with filtering", "Working", "Done", "LogsView + LogsPoller components"),
    ("System Information", "Log Level Runtime Change", "Change log level without restart", "Working", "Done", "handleSetLogLevel with API call"),
    ("System Information", "Clear Logs", "Clear system log buffer", "Working", "Done", "clearLogs function"),
    ("System Information", "System Restart", "Restart LightNVR service", "Working", "Done", "restartSystem function"),
    ("System Information", "System Shutdown", "Shutdown LightNVR service", "Working", "Done", "shutdownSystem function"),
    ("System Information", "Health Check API", "System health endpoint", "Working", "Done", "api_handlers_health.c (17KB)"),

    # ─── Retention & Storage ─────────────────────────────────────────
    ("Retention & Storage", "Global Retention Policy", "Days-based retention for all recordings", "Working", "Done", "retention_days in settings"),
    ("Retention & Storage", "Per-Stream Retention", "Stream-specific retention rules", "Working", "Done", "api_handlers_retention.c"),
    ("Retention & Storage", "Disk Space Management", "Monitor and manage storage capacity", "Working", "Done", "storage_manager.c"),
    ("Retention & Storage", "Auto Cleanup", "Automatic deletion of expired recordings", "Working", "Done", "auto_delete_oldest mechanism"),
    ("Retention & Storage", "Protected Recordings", "Exempt recordings from auto-deletion", "Working", "Done", "Protection toggle in UI"),
    ("Retention & Storage", "Storage Path Config", "Configurable recording storage location", "Working", "Done", "storage_path setting"),

    # ─── Recording Engine ────────────────────────────────────────────
    ("Recording Engine", "MP4 Recording", "Direct MP4 container writing", "Working", "Done", "mp4_recording_core.c + mp4_writer.c"),
    ("Recording Engine", "MP4 Segment Recording", "Segmented MP4 for continuous recording", "Working", "Done", "mp4_segment_recorder.c (69KB)"),
    ("Recording Engine", "HLS Recording/Streaming", "HLS segment generation", "Working", "Done", "hls_writer.c + hls_streaming.c"),
    ("Recording Engine", "Audio Transcoding (AAC)", "Transcode incompatible audio to AAC for MP4", "Working", "Done", "Fixed G.711 μ-law → AAC transcoding"),
    ("Recording Engine", "FFmpeg Utilities", "Codec/format handling via FFmpeg", "Working", "Done", "ffmpeg_utils.c (42KB)"),
    ("Recording Engine", "Timestamp Management", "Accurate timestamp handling across recordings", "Working", "Done", "timestamp_manager.c (25KB)"),
    ("Recording Engine", "Packet Buffering", "Efficient packet buffer for stream processing", "Working", "Done", "packet_buffer.c"),

    # ─── Core Infrastructure ─────────────────────────────────────────
    ("Core Infrastructure", "Stream Manager", "Centralized stream lifecycle management", "Working", "Done", "stream_manager.c (32KB)"),
    ("Core Infrastructure", "Stream State Machine", "State transitions for stream lifecycle", "Working", "Done", "stream_state.c (30KB)"),
    ("Core Infrastructure", "Configuration Manager", "INI-based config with validation", "Working", "Done", "config.c (52KB)"),
    ("Core Infrastructure", "Daemon Mode", "Run as background system service", "Working", "Done", "daemon.c with systemd integration"),
    ("Core Infrastructure", "Logger System", "File + syslog + JSON logging", "Working", "Done", "logger.c + logger_json.c"),
    ("Core Infrastructure", "Shutdown Coordinator", "Graceful shutdown with resource cleanup", "Working", "Done", "shutdown_coordinator.c"),
    ("Core Infrastructure", "go2rtc Integration", "Managing go2rtc process lifecycle", "Working", "Done", "go2rtc directory with 8 source files"),
    ("Core Infrastructure", "Database Migrations", "Automatic schema upgrades", "Working", "Done", "db_migrations.c + sqlite_migrate.c"),
    ("Core Infrastructure", "Database Backup", "SQLite database backup", "Working", "Done", "db_backup.c"),
    ("Core Infrastructure", "Thread Management", "Multi-threaded worker pool", "Working", "Done", "thread_utils.c + mongoose_server_multithreading.c"),
    ("Core Infrastructure", "Mongoose Web Server", "Embedded HTTP server", "Working", "Done", "mongoose_server.c (47KB)"),
    ("Core Infrastructure", "Rate Limiting", "API request rate limiting", "Working", "Done", "Documented in API.md"),

    # ─── Docker & Deployment ─────────────────────────────────────────
    ("Docker & Deployment", "Dockerfile", "Multi-stage build for production", "Working", "Done", "Dockerfile (9KB)"),
    ("Docker & Deployment", "Dockerfile Alpine", "Minimal Alpine-based image", "Working", "Done", "Dockerfile.alpine"),
    ("Docker & Deployment", "Docker Compose", "Full stack deployment config", "Working", "Done", "docker-compose.yml"),
    ("Docker & Deployment", "Docker Entrypoint", "Auto-init config, DB, web assets", "Working", "Done", "docker-entrypoint.sh (9KB)"),
    ("Docker & Deployment", "Volume Persistence", "Config + data volume separation", "Working", "Done", "/etc/lightnvr + /var/lib/lightnvr/data"),
    ("Docker & Deployment", "Environment Variables", "TZ, GO2RTC_CONFIG_PERSIST, LIGHTNVR_AUTO_INIT", "Working", "Done", "3 env vars documented"),
    ("Docker & Deployment", "Port Mapping", "8080 (web), 8554 (RTSP), 8555 (WebRTC), 1984 (go2rtc)", "Working", "Done", "4 exposed ports"),
    ("Docker & Deployment", "Build Scripts", "Build automation with release mode", "Working", "Done", "scripts/build.sh + install.sh"),

    # ─── API Endpoints ───────────────────────────────────────────────
    ("REST API", "GET /api/streams", "List all configured streams", "Working", "Done", "JSON response with stream array"),
    ("REST API", "GET /api/streams/{id}", "Get specific stream details", "Working", "Done", "Single stream object"),
    ("REST API", "POST /api/streams", "Add new stream", "Working", "Done", "Create with full config"),
    ("REST API", "PUT /api/streams/{id}", "Update existing stream", "Working", "Done", "Full update support"),
    ("REST API", "DELETE /api/streams/{id}", "Remove stream", "Working", "Done", "Cascade delete"),
    ("REST API", "GET /api/recordings", "List recordings with pagination/filters", "Working", "Done", "Paginated response"),
    ("REST API", "GET /api/recordings/{id}", "Get specific recording", "Working", "Done", "Single recording object"),
    ("REST API", "DELETE /api/recordings/{id}", "Delete recording", "Working", "Done", "File + DB cleanup"),
    ("REST API", "GET /api/settings", "Get system settings", "Working", "Done", "Full settings object"),
    ("REST API", "POST /api/settings", "Update system settings", "Working", "Done", "Partial update support"),
    ("REST API", "GET /api/system", "System info (CPU, RAM, storage)", "Working", "Done", "Real-time stats"),
    ("REST API", "GET /api/system/logs", "Fetch system logs", "Working", "Done", "With level filtering"),
    ("REST API", "Streaming APIs (HLS/MJPEG)", "Live stream endpoints", "Working", "Done", "HLS playlist + MJPEG stream"),
    ("REST API", "PTZ Control APIs", "Move, stop, home, presets, capabilities", "Working", "Done", "5 PTZ endpoints"),
    ("REST API", "Detection APIs", "Results, models, zones", "Working", "Done", "3 detection endpoint groups"),
    ("REST API", "go2rtc Proxy APIs", "Proxy requests to go2rtc", "Working", "Done", "api_handlers_go2rtc_proxy.c (24KB)"),
    ("REST API", "Timeline APIs", "Timeline data per stream/date", "Working", "Done", "api_handlers_timeline.c (30KB)"),
    ("REST API", "User Management APIs", "CRUD + API keys", "Working", "Done", "api_handlers_users.c (38KB)"),
    ("REST API", "Retention Policy APIs", "Per-stream retention config", "Working", "Done", "api_handlers_retention.c"),
    ("REST API", "ONVIF Discovery APIs", "Discover and query ONVIF devices", "Working", "Done", "api_handlers_onvif.c (20KB)"),
    ("REST API", "Health Check API", "Liveness/readiness checks", "Working", "Done", "api_handlers_health.c"),
    ("REST API", "Batch Operations APIs", "Batch delete with progress tracking", "Working", "Done", "api_handlers_recordings_batch.c"),

    # ─── Frontend Architecture ───────────────────────────────────────
    ("Frontend Architecture", "Preact SPA Framework", "Lightweight 3KB React alternative", "Working", "Done", "Preact + JSX components"),
    ("Frontend Architecture", "Tailwind CSS Styling", "Utility-first CSS framework", "Working", "Done", "tailwind.config.js with custom theme"),
    ("Frontend Architecture", "Vite Build System", "Fast dev server and bundler", "Working", "Done", "vite.config.js with multi-page setup"),
    ("Frontend Architecture", "Query Client", "Custom data fetching with caching", "Working", "Done", "query-client.js (7KB)"),
    ("Frontend Architecture", "Fetch Utilities", "Centralized API helpers with retry/timeout", "Working", "Done", "fetch-utils.js (8KB)"),
    ("Frontend Architecture", "Toast Notifications", "User feedback system", "Working", "Done", "Toast + ToastContainer components"),
    ("Frontend Architecture", "Loading Indicators", "Content/page loading states", "Working", "Done", "LoadingIndicator + ContentLoader"),
    ("Frontend Architecture", "UI Modal System", "Reusable modal infrastructure", "Working", "Done", "UI.jsx (32KB) with modal helpers"),
    ("Frontend Architecture", "Responsive Design", "Mobile-friendly layouts", "Working", "Done", "Grid + responsive breakpoints"),
    ("Frontend Architecture", "Test Suite", "Jest-based frontend tests", "Working", "Done", "web/tests/ with 12 test files"),

    # ─── Playwright Testing ──────────────────────────────────────────
    ("Testing & Docs", "Playwright Config", "E2E test configuration", "Working", "Done", "playwright.config.ts"),
    ("Testing & Docs", "Screenshot Automation", "Automated documentation screenshots", "Working", "Done", "scripts/update-documentation-media.sh"),
    ("Testing & Docs", "Theme Screenshot Variants", "All theme screenshots for docs", "Working", "Done", "--all-themes flag"),
    ("Testing & Docs", "Stress Testing", "Load/stress test scripts", "Working", "Done", "stress_test.sh"),
    ("Testing & Docs", "Backend Test Suite", "C-based test suite", "Working", "Done", "tests/ directory with 33 files"),
    ("Testing & Docs", "Comprehensive Documentation", "API, architecture, build, config, troubleshooting docs", "Working", "Done", "docs/ with 32 markdown files"),
]


def create_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "LightNVR Feature Analysis"

    # ── Column widths ───────────────────────────────────────────────
    col_widths = [8, 28, 50, 16, 16, 50]
    headers = ["#", "Feature", "Sub-Feature / Detail", "Status", "Test Status", "Notes / Implementation"]

    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Title row ───────────────────────────────────────────────────
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = "LightNVR – Full Feature Analysis Report"
    title_cell.font = Font(name="Calibri", bold=True, color="FFFFFF", size=16)
    title_cell.fill = PatternFill(start_color="111827", end_color="111827", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    # ── Subtitle row ────────────────────────────────────────────────
    ws.merge_cells("A2:F2")
    sub_cell = ws["A2"]
    sub_cell.value = "Generated: 2026-02-11  |  Project: opensensor/lightNVR  |  Total Features: {}".format(len(FEATURES))
    sub_cell.font = Font(name="Calibri", italic=True, color="9CA3AF", size=10)
    sub_cell.fill = PatternFill(start_color="111827", end_color="111827", fill_type="solid")
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 25

    # ── Legend row ──────────────────────────────────────────────────
    ws.merge_cells("A3:F3")
    legend_cell = ws["A3"]
    legend_cell.value = "🟢 Working/Done = Green  |  🔴 Not Working/Issue = Red  |  🟡 Testing/Partial = Yellow  |  🔵 Needs Review = Blue"
    legend_cell.font = Font(name="Calibri", size=10, color="4B5563")
    legend_cell.fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    legend_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 25

    # ── Header row ──────────────────────────────────────────────────
    header_row = 4
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGNMENT_CENTER
        cell.border = THIN_BORDER
    ws.row_dimensions[header_row].height = 30

    # ── Data rows ───────────────────────────────────────────────────
    current_row = header_row + 1
    current_category = None
    feature_num = 0
    data_rows = []  # Track rows that have Status/Test Status cells for dropdowns

    for category, feature, detail, status, test_status, notes in FEATURES:
        # Section header when category changes
        if category != current_category:
            current_category = category
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
            cell = ws.cell(row=current_row, column=1, value=f"▶  {category}")
            cell.font = FONT_SECTION
            cell.fill = FILL_SECTION
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = THIN_BORDER
            ws.row_dimensions[current_row].height = 28
            current_row += 1

        feature_num += 1
        alt_fill = FILL_GRAY if feature_num % 2 == 0 else FILL_WHITE

        # Column A: #
        cell = ws.cell(row=current_row, column=1, value=feature_num)
        cell.font = FONT_NORMAL
        cell.fill = alt_fill
        cell.alignment = ALIGNMENT_CENTER
        cell.border = THIN_BORDER

        # Column B: Feature
        cell = ws.cell(row=current_row, column=2, value=feature)
        cell.font = FONT_BOLD
        cell.fill = alt_fill
        cell.alignment = ALIGNMENT_LEFT
        cell.border = THIN_BORDER

        # Column C: Detail
        cell = ws.cell(row=current_row, column=3, value=detail)
        cell.font = FONT_NORMAL
        cell.fill = alt_fill
        cell.alignment = ALIGNMENT_LEFT
        cell.border = THIN_BORDER

        # Column D: Status (color driven by conditional formatting)
        cell = ws.cell(row=current_row, column=4, value=status)
        cell.font = FONT_NORMAL
        cell.fill = alt_fill
        cell.alignment = ALIGNMENT_CENTER
        cell.border = THIN_BORDER

        # Column E: Test Status (color driven by conditional formatting)
        cell = ws.cell(row=current_row, column=5, value=test_status)
        cell.font = FONT_NORMAL
        cell.fill = alt_fill
        cell.alignment = ALIGNMENT_CENTER
        cell.border = THIN_BORDER

        # Column F: Notes
        cell = ws.cell(row=current_row, column=6, value=notes)
        cell.font = FONT_NORMAL
        cell.fill = alt_fill
        cell.alignment = ALIGNMENT_LEFT
        cell.border = THIN_BORDER

        ws.row_dimensions[current_row].height = 22
        data_rows.append(current_row)
        current_row += 1

    # ── Summary section ─────────────────────────────────────────────
    current_row += 1
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
    cell = ws.cell(row=current_row, column=1, value="📊  Summary Statistics")
    cell.font = Font(name="Calibri", bold=True, color="FFFFFF", size=13)
    cell.fill = PatternFill(start_color="111827", end_color="111827", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[current_row].height = 32
    current_row += 1

    # Count categories
    categories = {}
    for cat, *_ in FEATURES:
        categories[cat] = categories.get(cat, 0) + 1

    # Count statuses
    status_counts = {}
    for *_, status, test_status, _ in FEATURES:
        status_counts[status] = status_counts.get(status, 0) + 1

    test_counts = {}
    for *_, test_status, _ in FEATURES:
        test_counts[test_status] = test_counts.get(test_status, 0) + 1

    summary_data = [
        ("Total Features", str(len(FEATURES))),
        ("Total Categories", str(len(categories))),
        ("", ""),
        ("Status Breakdown:", ""),
    ]
    for s, c in sorted(status_counts.items()):
        summary_data.append((f"  {s}", str(c)))

    summary_data.append(("", ""))
    summary_data.append(("Test Status Breakdown:", ""))
    for s, c in sorted(test_counts.items()):
        summary_data.append((f"  {s}", str(c)))

    summary_data.append(("", ""))
    summary_data.append(("Category Breakdown:", ""))
    for cat, count in sorted(categories.items()):
        summary_data.append((f"  {cat}", str(count)))

    for label, value in summary_data:
        cell_a = ws.cell(row=current_row, column=1)
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
        cell_a.value = label
        cell_a.font = FONT_BOLD if not label.startswith("  ") else FONT_NORMAL
        cell_a.fill = FILL_WHITE
        cell_a.alignment = ALIGNMENT_LEFT
        cell_a.border = THIN_BORDER

        cell_b = ws.cell(row=current_row, column=5)
        ws.merge_cells(start_row=current_row, start_column=5, end_row=current_row, end_column=6)
        cell_b.value = value
        cell_b.font = FONT_BOLD
        cell_b.fill = FILL_WHITE
        cell_b.alignment = ALIGNMENT_CENTER
        cell_b.border = THIN_BORDER

        current_row += 1

    # ── Dropdown data validation for Status (col D) ─────────────────
    status_dv = DataValidation(
        type="list",
        formula1='"Working,Not Working,Issue,Partial,Needs Review"',
        allow_blank=False,
        showDropDown=False,  # False = SHOW the dropdown arrow in Excel
    )
    status_dv.error = "Please select a valid status"
    status_dv.errorTitle = "Invalid Status"
    status_dv.prompt = "Select feature status"
    status_dv.promptTitle = "Status"
    ws.add_data_validation(status_dv)

    # ── Dropdown data validation for Test Status (col E) ────────────
    test_dv = DataValidation(
        type="list",
        formula1='"Done,Testing,Needs Review,Not Started"',
        allow_blank=False,
        showDropDown=False,
    )
    test_dv.error = "Please select a valid test status"
    test_dv.errorTitle = "Invalid Test Status"
    test_dv.prompt = "Select test status"
    test_dv.promptTitle = "Test Status"
    ws.add_data_validation(test_dv)

    # Apply dropdowns to all feature data rows
    for row_num in data_rows:
        status_dv.add(ws.cell(row=row_num, column=4))   # Column D
        test_dv.add(ws.cell(row=row_num, column=5))     # Column E

    # ── Conditional formatting for auto-coloring on selection ────────
    # Define color rules: (value, fill_color, font_color)
    status_rules = [
        ("Working",      "D1FAE5", "065F46"),  # Green bg, dark green text
        ("Not Working",  "FEE2E2", "991B1B"),  # Red bg, dark red text
        ("Issue",        "FEE2E2", "991B1B"),  # Red bg, dark red text
        ("Partial",      "FEF3C7", "92400E"),  # Yellow bg, dark amber text
        ("Needs Review", "DBEAFE", "1E40AF"),  # Blue bg, dark blue text
    ]
    test_rules = [
        ("Done",         "D1FAE5", "065F46"),  # Green bg, dark green text
        ("Testing",      "FEF3C7", "92400E"),  # Yellow bg, dark amber text
        ("Needs Review", "DBEAFE", "1E40AF"),  # Blue bg, dark blue text
        ("Not Started",  "F3F4F6", "6B7280"),  # Gray bg, gray text
    ]

    first_data_row = data_rows[0]
    last_data_row = data_rows[-1]
    col_d_range = f"D{first_data_row}:D{last_data_row}"
    col_e_range = f"E{first_data_row}:E{last_data_row}"

    for value, bg_color, font_color in status_rules:
        ws.conditional_formatting.add(
            col_d_range,
            CellIsRule(
                operator="equal",
                formula=[f'"{value}"'],
                fill=PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid"),
                font=Font(name="Calibri", bold=True, size=10, color=font_color),
            )
        )

    for value, bg_color, font_color in test_rules:
        ws.conditional_formatting.add(
            col_e_range,
            CellIsRule(
                operator="equal",
                formula=[f'"{value}"'],
                fill=PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid"),
                font=Font(name="Calibri", bold=True, size=10, color=font_color),
            )
        )

    # ── Freeze panes ────────────────────────────────────────────────
    ws.freeze_panes = "A5"

    # ── Auto-filter ─────────────────────────────────────────────────
    ws.auto_filter.ref = f"A4:F{header_row + len(FEATURES) + len(categories)}"

    # ── Save ────────────────────────────────────────────────────────
    output_path = "/home/rahad/work/lightNVR/LightNVR_Feature_Analysis.xlsx"
    wb.save(output_path)
    print(f"✅ Excel sheet saved to: {output_path}")
    print(f"   Total features: {len(FEATURES)}")
    print(f"   Total categories: {len(categories)}")


if __name__ == "__main__":
    create_excel()
